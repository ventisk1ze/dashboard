from flask import Flask, render_template, request, redirect, url_for, flash
from flask.globals import session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from sqlalchemy.sql.expression import extract
from werkzeug.utils import secure_filename
import os
import openpyxl as opl
import datetime as dt
from dateutil.relativedelta import relativedelta

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///test.db'
db = SQLAlchemy(app)

class Debt(db.Model):
    id = db.Column(db.Integer, primary_key = True)
    date = db.Column(db.DateTime)
    amount = db.Column(db.Float)
    employee = db.Column(db.String(100))

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}
UPLOAD_FOLDER = './'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
def save_data(filename):
    workbook = opl.load_workbook(filename)
    worksheet = workbook.get_sheet_by_name('Данные')
    for i in range(2, worksheet.max_row):
        date = worksheet.cell(row = i, column = 1).value
        amount = worksheet.cell(row = i, column = 2).value
        employee = worksheet.cell(row = i, column = 3).value
        entry = Debt(date = date, amount = amount, employee = employee)
        db.session.add(entry)
    db.session.commit()
def debt_sum():
    result = db.session.query(func.sum(Debt.amount))
    return result.scalar()
def overdue_debt():
    yearAgo = dt.datetime.now() - relativedelta(years=1)
    result = db.session.query(func.sum(Debt.amount)).filter(Debt.date <= yearAgo).scalar()
    return result
@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('Файл не найден')
            return redirect(request.url)
        file = request.files['file']
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            flash('Файл не выбран')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            save_data(filename)
            return redirect(url_for('upload_file', name=filename))
    entries = Debt.query.all()
    return render_template('index.html', entries = entries, ds = debt_sum(), od = overdue_debt())

if __name__ == "__main__":
    app.run(debug = True)
